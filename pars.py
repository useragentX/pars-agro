import os
import re
import openpyxl
import requests
from bs4 import BeautifulSoup as bs

URL = 'https://agro-market24.ru'
URL_TEMPLATE = 'https://agro-market24.ru/catalog/'

headers = requests.utils.default_headers()
headers.update(
    {
        'User-Agent': 'My User Agent 1.0',
    }
)

if not os.path.isdir("result"):
    os.mkdir("result")

hs_category = []
hs_product = ['https://agro-market24.ru/catalog/item/standartnaya_karta_zakrytogo_kluba_lux/',
'https://agro-market24.ru/catalog/item/ukorenitel_kornevin_krepkaya_osnova_universalnyy_tm_zolotoy_urozhay_100g/',
'https://agro-market24.ru/catalog/item/stimulyator_rosta_gumat_osnova_zdorovya_universalnyy_tm_zolotoy_urozhay_100g/',
'https://agro-market24.ru/catalog/item/universalnoe_pitanie_udobrenie_yarkie_kraski_dlya_tsvetushchikh_tm_zolotoy_urozhay_100g/'
]

def category_up(url, dir, lvl):
    r = requests.get(url, headers=headers)
    soup = bs(r.text, "lxml")
    section = soup.find('div', class_='section-list')
    if section:
        section_list_links = section.find_all('a', class_='section-item')
        for i in section_list_links:
            section_name = i.find('span', class_='section-item-name')
            if section_name in hs_category:
                break
            else:
                hs_category.append(section_name)

            dir_tree = f"{dir}/{section_name.get_text(strip=True)}"
            if not os.path.isdir(dir_tree):
                os.makedirs(dir_tree)
            print(f"{lvl}/ {section_name.get_text(strip=True)}")
            category_up(URL+i['href'], dir_tree, " "+lvl+"-")

    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Название"
        ws['B1'] = "Артикул"
        ws['C1'] = "Цена"
        ws['D1'] = "Изображение"
        ws['E1'] = "Все картинки"
        ws['F1'] = "Ссылка на товар"


        pp_title = []
        pp_items = []
        page_count = 1
        max_page_count = 1
        max_page_count = 1
        product_count = 2
        while True:
            r = requests.get(f'{url}?PAGEN_1={page_count}', headers=headers)
            if page_count == 1:
                max_page_count_re = re.search(r"NavPageCount':[0-9]+,", r.text)[0]
                max_page_count = int(re.search(r"[0-9]+", max_page_count_re)[0])

            
            soup = bs(r.text, "lxml")
            products = soup.find('div', class_='links')
            product_list_name = products.find_all('a')
            for i in product_list_name:
                if URL+i['href'] in hs_product:
                    continue

                product_read = requests.get(URL+i['href'], headers=headers)
                product_soup = bs(product_read.text, "lxml")

                product_articul = product_soup.find('div', class_='articul')
                if product_articul:
                    product_articul = int(product_articul.find('span', class_='value').get_text(strip=True))
                else:
                    product_articul = product_soup.find('div', class_='artnumber').get_text(strip=True)
                    product_articul = int(re.search(r"[0-9]+", product_articul)[0])


                product_title = product_soup.find('h1', class_='title')
                if product_title:
                    product_title = product_title.get_text(strip=True)
                else:
                    product_title = product_soup.find('div', class_='name')
                    product_title = product_title.find('h1').get_text(strip=True)


                product_price = product_soup.find(class_='price')
                if product_price:
                    product_price = product_price.find(text=True).get_text(strip=True)

                product_img = ""
                try:
                    imgs = product_soup.find_all('img', class_='slide-img')
                    if imgs:
                        p = requests.get(URL+imgs[0]['src'], headers=headers)
                        out = open(f"img/{product_articul}.jpg", "wb")
                        out.write(p.content)
                        out.close()
                        main_img = openpyxl.drawing.image.Image(f"img/{product_articul}.jpg")
                        main_img.width = 50
                        main_img.height = 80
                        ws.column_dimensions['D'].width = 20
                        ws.row_dimensions[product_count].height = 60
                        ws.add_image(main_img, f"D{product_count}")

                        for img in imgs:
                            product_img += f"{URL}{img['src']} \n"
                    else:
                        product_img = product_soup.find('div', class_='pict')
                        product_img = URL + product_img['href']
                except:
                    pass
                
                product_property_text = ""
                product_property = product_soup.find_all('div', class_='props-row')
                if product_property:
                    pp_dict = {}
                    for property in product_property:
                        property_title = property.find('span', class_='props-name')
                        property_value = property.find('span', class_='props-value')
                        
                        if property_title and property_value:
                            property_title = property_title.get_text(strip=True)
                            property_value = property_value.get_text(strip=True)
                            if not (property_title in pp_title):
                                pp_title.append(property_title)
                            pp_dict[property_title] = property_value

                    pp_items.append(pp_dict)
                else:
                    pp_items.append({})

                ws[f'A{product_count}'] = product_title
                ws[f'B{product_count}'] = product_articul
                ws[f'C{product_count}'] = product_price

                ws[f'E{product_count}'] = product_img
                ws[f'F{product_count}'] = URL+i['href']

                product_count += 1


            if page_count != max_page_count:
                
                page_count += 1
            else:
                page_count = 1
                break

        pp_i = 0
        for pp in pp_title:
            ws.cell(row=1, column=7+pp_i).value = pp
            pp_i += 1

        pp_item_i = 2
        for pp_item in pp_items:
            for key, value in pp_item.items():
                if key in pp_title:
                    ws.cell(row=pp_item_i, column=7+pp_title.index(key)).value = value

            pp_item_i += 1

        '''
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                         max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        '''


        file_title = soup.find('h1', class_='title').get_text(strip=True) 
        wb.save(f'{dir}/{file_title}.xlsx')

category_up(URL_TEMPLATE, "result", "-")